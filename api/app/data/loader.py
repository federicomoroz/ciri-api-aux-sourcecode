"""
Excel → SQLite loader.

Dataset quirks:
- Row 1: decorative merged title (skip)
- Row 2: real headers (used as column mapping)
- Data starts at Row 3
- Sheet names contain emojis → use substring matching
- All dates/timestamps are strings (openpyxl reads them as str)
- Logs.Codigo is a string, not int
- Policies "Politica" column: "POL-XXX-NNN — Description" (em-dash)
"""

import logging
import sqlite3
from datetime import UTC, datetime

import openpyxl

from ..domain.constants import POLICY_SEED_BLOQUEANTES, POLICY_SEED_SLA_DIAS
from . import esquema

logger = logging.getLogger(__name__)


def _find_sheet(wb: openpyxl.Workbook, keyword: str) -> str:
    """Find sheet by case-insensitive keyword substring (handles emoji prefixes)."""
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return name
    raise ValueError(f"Sheet containing '{keyword}' not found. Available: {wb.sheetnames}")


def _parse_policy_field(raw: str) -> tuple[str, str]:
    """'POL-FRD-001 — Score minimo' → ('POL-FRD-001', 'Score minimo')"""
    if not raw:
        return ("", "")
    # Try em-dash with spaces (U+2014)
    if " \u2014 " in raw:
        parts = raw.split(" \u2014 ", maxsplit=1)
        return parts[0].strip(), parts[1].strip()
    # Try regular dash with spaces
    if " - " in raw:
        parts = raw.split(" - ", maxsplit=1)
        return parts[0].strip(), parts[1].strip()
    # Fallback: treat whole thing as code
    return raw.strip(), ""


def _parse_sheet(wb: openpyxl.Workbook, keyword: str, columns: list[str]) -> list[dict]:
    """Generic parser: find sheet by keyword, skip row 1, row 2=headers (ignored), data from row 3."""
    sheet_name = _find_sheet(wb, keyword)
    ws = wb[sheet_name]
    rows = []
    for row_num in range(3, ws.max_row + 1):
        values = [ws.cell(row=row_num, column=c).value for c in range(1, len(columns) + 1)]
        if all(v is None for v in values):
            continue
        row_dict = dict(zip(columns, values, strict=True))

        # Type coercions
        if "amount_usd" in row_dict and row_dict["amount_usd"] is not None:
            row_dict["amount_usd"] = float(row_dict["amount_usd"])
        if "fraud_score" in row_dict and row_dict["fraud_score"] is not None:
            row_dict["fraud_score"] = int(row_dict["fraud_score"])
        if "resolution_days" in row_dict and row_dict["resolution_days"] is not None:
            row_dict["resolution_days"] = int(row_dict["resolution_days"])
        # Ensure date/timestamp fields are strings
        for date_field in ("date", "open_date", "close_date", "timestamp"):
            if date_field in row_dict and row_dict[date_field] is not None:
                row_dict[date_field] = str(row_dict[date_field])
        # HTTP code must be string
        if "code" in row_dict and row_dict["code"] is not None:
            try:
                row_dict["code"] = str(int(float(str(row_dict["code"])))) if row_dict["code"] != "" else "0"
            except (ValueError, TypeError):
                row_dict["code"] = str(row_dict["code"])

        rows.append(row_dict)
    return rows


def load_excel(file_path: str) -> dict:
    """Load all 4 data sheets from the Excel file (skips README sheet).

    Returns:
        {
            'transactions': [...],  # 100 rows
            'cases': [...],         # 60 rows
            'policies': [...],      # 17 rows
            'logs': [...],          # 150 rows
        }
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        logger.error("Failed to load Excel file: %s", file_path, exc_info=True)
        raise

    transactions = _parse_sheet(wb, "Transacciones", [
        "id", "client_id", "merchant", "amount_usd", "date",
        "payment_method", "country", "channel", "device",
        "fraud_score", "status", "notes",
    ])

    raw_cases = _parse_sheet(wb, "Hist", [
        "case_id", "transaction_id", "motivo", "resolution",
        "resolution_days", "analyst", "observations",
        "open_date", "close_date",
    ])

    raw_policies = _parse_sheet(wb, "Pol", [
        "category", "politica_raw", "description", "reference",
    ])

    logs = _parse_sheet(wb, "Logs", [
        "timestamp", "transaction_id", "event", "service",
        "code", "detail", "severity",
    ])

    # Parse policy code + name from compound "Politica" field
    policies = []
    for p in raw_policies:
        code, name = _parse_policy_field(p.get("politica_raw") or "")
        policies.append({
            "code": code,
            "name": name,
            "category": p.get("category", ""),
            "description": p.get("description", ""),
            "reference": p.get("reference", ""),
        })

    return {
        "transactions": transactions,
        "cases": raw_cases,
        "policies": policies,
        "logs": logs,
    }


def _crear_esquema(conn: sqlite3.Connection) -> None:
    """El esquema sale de `data/esquema.py`, que es su unico dueno.

    Este modulo declaraba cinco tablas por su cuenta mientras `esquema.py`
    declaraba otras tres y ademas le agregaba columnas a `policies`: la misma
    tabla se creaba en un archivo y se migraba en otro, asi que responder que
    forma tiene la base pedia leer los dos y saber cual corre primero.
    """
    for ddl in esquema.TODO:
        conn.execute(ddl)


def _insertar(conn: sqlite3.Connection, data: dict) -> None:
    """Vuelca el Excel en las tablas.

    Transacciones, casos y politicas se deduplican por su clave de negocio. Los
    logs no tienen una, asi que se reemplazan enteros: son datos de referencia
    que solo salen del Excel y nadie escribe en runtime. Reinsertarlos en cada
    corrida del seed duplicaria las 150 filas y falsearia la deteccion de
    patrones, que cuenta eventos repetidos.
    """
    now = datetime.now(UTC).isoformat()

    conn.executemany(
        "INSERT OR IGNORE INTO transactions VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        [(
            t["id"], t["client_id"], t["merchant"], t["amount_usd"], t["date"],
            t["payment_method"], t["country"], t["channel"], t["device"],
            t["fraud_score"], t["status"], t["notes"],
        ) for t in data["transactions"]],
    )

    conn.executemany(
        "INSERT OR IGNORE INTO cases VALUES (?,?,?,?,?,?,?,?,?)",
        [(
            c["case_id"], c["transaction_id"], c["motivo"], c["resolution"],
            c["resolution_days"], c["analyst"], c["observations"],
            c["open_date"], c["close_date"],
        ) for c in data["cases"]],
    )

    # Columnas explicitas: con VALUES posicional, agregar una columna rompe la
    # carga en silencio o la desplaza.
    conn.executemany(
        "INSERT OR IGNORE INTO policies "
        "(code, name, category, description, reference, created_at, updated_at, "
        " puede_bloquear, sla_dias) VALUES (?,?,?,?,?,?,?,?,?)",
        [(
            p["code"], p["name"], p["category"],
            p["description"], p["reference"], now, now,
            int(p["code"] in POLICY_SEED_BLOQUEANTES),
            POLICY_SEED_SLA_DIAS.get(p["code"]),
        ) for p in data["policies"]],
    )

    conn.execute("DELETE FROM logs")
    conn.executemany(
        "INSERT INTO logs (timestamp, transaction_id, event, service, code, detail, severity) VALUES (?,?,?,?,?,?,?)",
        [(
            log["timestamp"], log["transaction_id"], log["event"],
            log["service"], log["code"], log["detail"], log["severity"],
        ) for log in data["logs"]],
    )


def init_sqlite(db_path: str, data: dict) -> None:
    """Crea el esquema y carga los datos.

    Idempotente: correr el seed dos veces deja la base igual que correrlo una.
    """
    conn = sqlite3.connect(db_path)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        _crear_esquema(conn)
        _insertar(conn, data)
        conn.commit()
    finally:
        conn.close()
